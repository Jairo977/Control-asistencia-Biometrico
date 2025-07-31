from django.db import connections
def obtener_ultima_marcacion(id_usuario):
    from .db_utils import get_db_connection
    query = """
        SELECT TOP 1
            CAST(t.punch_time AS DATE) AS fecha,
            CAST(t.punch_time AS TIME(0)) AS hora
        FROM iclock_transaction t
        WHERE RTRIM(LTRIM(t.emp_code)) = RTRIM(LTRIM(?))
        ORDER BY t.punch_time DESC
    """
    conn = get_db_connection()
    if not conn:
        print(f"No se pudo conectar a la base de datos para usuario: {id_usuario}")
        return None
    try:
        with conn.cursor() as cur:
            cur.execute(query, [str(id_usuario).strip()])
            row = cur.fetchone()
            if row:
                return {'fecha': row.fecha, 'hora': row.hora}
            # Si no hay marcación, retorna None explícitamente
            return None
    except Exception as e:
        print(f"Error al obtener última marcación para usuario {id_usuario}: {e}")
    finally:
        conn.close()
    return None

def registros_hoy(request):
    from datetime import date
    registros = db_utils.obtener_reportes(
        fecha_inicio=date.today().strftime('%Y%m%d'),
        fecha_fin=date.today().strftime('%Y%m%d')
    )
    contexto = {
        'fecha': date.today().strftime('%d/%m/%Y'),
        'registros': registros,
        'titulo': 'Registros de Hoy',
        'total_atrasados': sum(1 for r in registros if r.get('observacion') == 'Atraso'),
        'no_han_marcado': sum(1 for r in registros if r.get('observacion') == 'Sin registro'),
    }
    return render(request, 'registros_hoy.html', contexto)

# Vista para el dashboard de registros hoy
def registros_hoy_dashboard(request):
    from datetime import date
    from .db_utils import get_db_connection

    # Consulta optimizada usando pyodbc directamente
    query = '''
        SELECT e.emp_code AS id_usuario, e.first_name AS nombre, d.dept_name AS departamento,
               MAX(CAST(t.punch_time AS DATE)) AS fecha,
               (SELECT TOP 1 CAST(t2.punch_time AS TIME(0))
                FROM iclock_transaction t2
                WHERE t2.emp_code = e.emp_code
                ORDER BY t2.punch_time DESC) AS hora
        FROM personnel_employee e
        JOIN personnel_department d ON e.department_id = d.id
        LEFT JOIN iclock_transaction t ON e.emp_code = t.emp_code
        GROUP BY e.emp_code, e.first_name, d.dept_name
        ORDER BY e.first_name
    '''
    usuarios = []
    conn = get_db_connection()
    if conn:
        try:
            with conn.cursor() as cur:
                cur.execute(query)
                rows = cur.fetchall()
                for idx, row in enumerate(rows, start=1):
                    fecha = row.fecha.strftime('%Y-%m-%d') if row.fecha else None
                    hora = str(row.hora) if row.hora else None
                    usuarios.append({
                        'item': idx,
                        'id_usuario': row.id_usuario,
                        'nombre': row.nombre,
                        'departamento': row.departamento,
                        'ultima_marcacion': {'fecha': fecha, 'hora': hora} if fecha and hora else None
                    })
        finally:
            conn.close()
    contexto = {
        'fecha': date.today().strftime('%d/%m/%Y'),
        'usuarios': usuarios,
        'total_usuarios': len(usuarios),
        'titulo': 'Lista de Usuarios',
    }
    return render(request, 'registros_hoy_dashboard.html', contexto)
def reportes_view(request):
    """
    Vista para la sección de reportes.
    Muestra información y acceso a reportes generados.
    """
    from datetime import date
    departamentos = db_utils.obtener_departamentos()
    stats = db_utils.obtener_estadisticas_dashboard()
    reportes = []
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    id_usuario = request.GET.get('id_usuario')
    nombre = request.GET.get('nombre')
    departamento = request.GET.get('departamento')
    # Si no hay fechas, usar el día actual en formato seguro
    if not fecha_inicio or not fecha_fin:
        hoy = date.today().strftime('%Y%m%d')
        fecha_inicio = fecha_inicio or hoy
        fecha_fin = fecha_fin or hoy
    else:
        # Si vienen en formato 'YYYY-MM-DD', convertir a 'YYYYMMDD'
        try:
            if '-' in fecha_inicio:
                fecha_inicio = date.fromisoformat(fecha_inicio).strftime('%Y%m%d')
            if '-' in fecha_fin:
                fecha_fin = date.fromisoformat(fecha_fin).strftime('%Y%m%d')
        except Exception:
            pass
    reportes = db_utils.obtener_reportes(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        id_usuario=id_usuario,
        nombre=nombre,
        departamento=departamento
    )
    contexto = {
        'titulo': 'Reportes de Asistencia',
        'descripcion': 'Aquí puedes consultar y descargar reportes generados en PDF y Excel.',
        'departamentos': departamentos,
        'reportes': reportes,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'id_usuario': id_usuario,
        'nombre': nombre,
        'departamento': departamento,
        'total_personal': stats.get('total_empleados', 0),
        'atrasados_hoy': stats.get('llegadas_tarde', 0),
        'no_han_marcado': stats.get('ausentes_hoy', 0),
    }
    return render(request, 'reportes_section.html', contexto)

def configuracion_view(request):
    """
    Vista para la sección de configuración del sistema.
    Permite ajustar parámetros y preferencias del sistema biométrico.
    """
    contexto = {
        'titulo': 'Configuración del Sistema',
        'descripcion': 'Ajusta parámetros, preferencias y opciones avanzadas del sistema biométrico.',
    }
    return render(request, 'configuracion_section.html', contexto)
def bienvenida(request):
    """
    Vista de bienvenida al sistema biométrico.
    Muestra información básica y guía de uso.
    """
    from datetime import date
    stats = db_utils.obtener_estadisticas_dashboard()
    contexto = {
        'titulo': 'Bienvenido al Sistema Biométrico',
        'descripcion': 'Este sistema permite gestionar y consultar la asistencia biométrica del personal, generar reportes y acceder a estadísticas en tiempo real.',
        'version': '1.0',
        'autor': 'Equipo de Desarrollo',
        'contacto': 'soporte@ejemplo.com',
        'fecha': date.today().strftime('%d/%m/%Y'),
        'registros_hoy': stats.get('empleados_registrados_hoy', 0),
        'total_atrasados': stats.get('llegadas_tarde', 0),
        'total_personal': stats.get('total_empleados', 0),
        'no_han_marcado': stats.get('ausentes_hoy', 0),
        'porcentaje_asistencia': stats.get('porcentaje_asistencia', 0.0),
        'promedio_horas': stats.get('promedio_horas', '0.0 hrs'),
    }
    return render(request, 'bienvenida.html', contexto)
"""
DJANGO VERSION - Refactorizada para usar db_utils
"""
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from datetime import datetime, date
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.utils import ImageReader
from io import BytesIO
import json
import textwrap

# Importar el módulo centralizado de base de datos
from . import db_utils 

# Verificar si openpyxl está disponible para las funciones de Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# --- FUNCIONES AUXILIARES ---

def draw_wrapped_text(canvas_obj, text, x, y, max_width, line_height=12, font_size=8):
    """
    Dibuja texto que se ajusta automáticamente en múltiples líneas
    """
    if not text:
        return y
    
    # Configurar fuente
    canvas_obj.setFont("Helvetica", font_size)
    
    # Estimar caracteres por línea basado en el ancho
    chars_per_line = max(1, int(max_width / (font_size * 0.6)))
    
    # Dividir el texto en líneas
    lines = textwrap.fill(str(text), width=chars_per_line).split('\n')
    
    # Dibujar cada línea
    current_y = y
    for line in lines:
        canvas_obj.drawString(x, current_y, line)
        current_y -= line_height
    
    return current_y - line_height  # Retornar la nueva posición Y


# --- VISTAS DE LA APLICACIÓN ---

@login_required
def index(request):
    """Renderiza el dashboard principal utilizando datos centralizados de db_utils."""
    try:
        stats = db_utils.obtener_estadisticas_dashboard()
        context = {
            'app_name': 'Sistema de Reportes Biométrico',
            'usuario': request.user.username,
            'rol': 'Administrador',
            'departamento': 'TI - Sistemas',
            'version': '2.1 Refactorizado',
            'fecha_actual': date.today().strftime('%d/%m/%Y'),
            'hora_actual': datetime.now().strftime('%H:%M:%S'),
            'total_usuarios': stats.get('total_empleados', 0),
            'presentes_hoy': stats.get('empleados_registrados_hoy', 0),
            'ausentes_hoy': stats.get('ausentes_hoy', 0),
            'llegadas_tarde': stats.get('llegadas_tarde', 0),
            'eficiencia_asistencia': stats.get('porcentaje_asistencia', 0.0),
            'promedio_horas_dia': stats.get('promedio_horas', '0.0 hrs'),
        }
    except Exception as e:
        print(f"Error obteniendo datos del dashboard: {e}")
        # Contexto de fallback en caso de error en la BD
        context = {
            'app_name': 'Sistema de Reportes Biométrico',
            'usuario': request.user.username,
            'rol': 'Administrador',
            'departamento': 'TI - Sistemas',
            'version': '2.1 Refactorizado (Error Mode)',
            'fecha_actual': date.today().strftime('%d/%m/%Y'),
            'hora_actual': datetime.now().strftime('%H:%M:%S'),
            'total_usuarios': 0, 'presentes_hoy': 0, 'ausentes_hoy': 0,
            'llegadas_tarde': 0, 'eficiencia_asistencia': 0.0, 'promedio_horas_dia': 'N/A',
        }
    return render(request, 'dashboard_ultra.html', context)

@login_required
@csrf_exempt
def api_reports(request):
    """API endpoint para obtener reportes de asistencia, usando db_utils."""
    try:
        # Extraer parámetros de la solicitud GET
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        id_usuario = request.GET.get('id_usuario')
        nombre = request.GET.get('nombre')
        departamento = request.GET.get('departamento')
        
        # Llamar a la función centralizada
        reportes = db_utils.obtener_reportes(fecha_inicio, fecha_fin, id_usuario, nombre, departamento)
        
        return JsonResponse({'success': True, 'data': reportes}, safe=False)
        
    except Exception as e:
        print(f"Error en api_reports: {e}")
        return JsonResponse({'success': False, 'error': str(e), 'data': []}, status=500)

@login_required
@csrf_exempt
def api_departamentos(request):
    """API endpoint para obtener la lista de departamentos, usando db_utils."""
    try:
        departamentos = db_utils.obtener_departamentos()
        return JsonResponse(departamentos, safe=False)
    except Exception as e:
        print(f"Error en api_departamentos: {e}")
        return JsonResponse([], safe=False, status=500)

@login_required
def api_empleados(request):
    """API endpoint para obtener empleados, usando db_utils."""
    try:
        departamento = request.GET.get('departamento', None)
        empleados = db_utils.obtener_empleados(departamento)
        return JsonResponse(empleados, safe=False)
    except Exception as e:
        print(f"Error en api_empleados: {e}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@csrf_exempt
def download_pdf(request):
    """Genera y descarga un reporte en PDF con texto ajustado automáticamente."""
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    id_usuario = request.GET.get('id_usuario')
    nombre = request.GET.get('nombre')
    departamento = request.GET.get('departamento')
    
    reportes = db_utils.obtener_reportes(fecha_inicio, fecha_fin, id_usuario, nombre, departamento)
    
    # Determinar el encabezado y nombre de archivo dinámicamente
    if id_usuario:
        nombre_empleado = reportes[0]['nombre'] if reportes and reportes[0]['nombre'] else f"ID_{id_usuario}"
        header_title = f"Reporte Individual de {nombre_empleado}"
        filename_pdf = f"Reporte_de_{nombre_empleado.replace(' ', '_')}.pdf"
    elif nombre:
        header_title = f"Reporte Individual de {nombre}"
        filename_pdf = f"Reporte_de_{nombre.replace(' ', '_')}.pdf"
    elif departamento:
        header_title = f"Reporte de {departamento}"
        filename_pdf = f"Reporte_de_{departamento.replace(' ', '_')}.pdf"
    else:
        header_title = "Reporte General"
        filename_pdf = "Reporte_General.pdf"

    buffer = BytesIO()
    # Usar landscape para todos los reportes para mejor aprovechamiento del espacio
    page_size = landscape(letter)
    p = canvas.Canvas(buffer, pagesize=page_size)
    width, height = page_size

    # Encabezado principal
    p.setFont("Helvetica-Bold", 16)
    p.drawString(40, height - 40, header_title.upper())
    p.setFont("Helvetica", 10)
    p.drawString(40, height - 60, f"Periodo: {fecha_inicio or ''} a {fecha_fin or ''}")
    p.drawString(40, height - 75, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    if not reportes:
        p.setFont("Helvetica-Bold", 14)
        p.drawString(40, height - 120, "NO HAY DATOS PARA MOSTRAR")
    else:
        # Detectar si es reporte individual
        es_reporte_individual = bool(id_usuario or nombre)
        
        if es_reporte_individual:
            # REPORTE INDIVIDUAL - Mostrar encabezado del empleado y solo fechas/marcaciones
            y = height - 100
            
            # Información del empleado
            if reportes:
                emp = reportes[0]
                
                # Encabezado de información del empleado
                p.setFont("Helvetica-Bold", 14)
                p.drawString(40, y, "INFORMACIÓN DEL EMPLEADO")
                y -= 30
                
                # Cédula
                p.setFont("Helvetica-Bold", 11)
                p.drawString(40, y, "Cédula:")
                p.setFont("Helvetica", 11)
                p.drawString(120, y, str(emp['id_usuario']) or 'No disponible')
                y -= 20
                
                # Nombre completo
                p.setFont("Helvetica-Bold", 11)
                p.drawString(40, y, "Nombre Completo:")
                p.setFont("Helvetica", 11)
                nombre_completo = emp['nombre'] or 'Sin nombre'
                # Ajustar texto largo para nombre
                if len(nombre_completo) > 50:
                    p.drawString(150, y, nombre_completo[:50])
                    y -= 12
                    p.drawString(150, y, nombre_completo[50:])
                else:
                    p.drawString(150, y, nombre_completo)
                y -= 20
                
                # Departamento
                p.setFont("Helvetica-Bold", 11)
                p.drawString(40, y, "Departamento:")
                p.setFont("Helvetica", 11)
                departamento_emp = emp['departamento'] or 'Sin departamento'
                # Ajustar texto largo para departamento
                if len(departamento_emp) > 45:
                    p.drawString(140, y, departamento_emp[:45])
                    y -= 12
                    p.drawString(140, y, departamento_emp[45:])
                else:
                    p.drawString(140, y, departamento_emp)
                y -= 35
                
                # Línea separadora gruesa
                p.setLineWidth(2)
                p.line(35, y, width-35, y)
                p.setLineWidth(1)
                y -= 20
            
            # Headers para reporte individual (sin ID, nombre, departamento)
            headers_individual = ["Fecha", "1ra Marcación", "2da Marcación", "3ra Marcación", "4ta Marcación", "Observaciones"]
            col_positions_individual = [40, 150, 270, 390, 510, 630]
            col_widths_individual = [100, 110, 110, 110, 110, 120]
            
            # Dibujar headers
            p.setFont("Helvetica-Bold", 10)
            for i, header in enumerate(headers_individual):
                p.drawString(col_positions_individual[i], y, header)
            
            # Línea separadora debajo de headers
            p.line(35, y-5, width-35, y-5)
            y -= 25
            
            # Datos de asistencia para reporte individual
            p.setFont("Helvetica", 9)
            base_row_height = 18
            
            for reporte in reportes:
                # Verificar si necesitamos nueva página
                if y < 80:
                    p.showPage()
                    y = height - 80
                    # Repetir headers en nueva página
                    p.setFont("Helvetica-Bold", 10)
                    for i, header in enumerate(headers_individual):
                        p.drawString(col_positions_individual[i], y, header)
                    p.line(35, y-5, width-35, y-5)
                    y -= 25
                    p.setFont("Helvetica", 9)
                
                # Resaltar atrasos
                if reporte.get('observacion') == 'Atraso':
                    p.setFillColorRGB(1, 0.9, 0.9)
                    p.rect(35, y-3, width-70, base_row_height-2, stroke=0, fill=1)
                    p.setFillColorRGB(0, 0, 0)
                
                # Solo datos relevantes para reporte individual
                datos_individual = [
                    reporte['fecha'] or "",
                    reporte['hora_ingreso'] or "",
                    reporte['hora_inicio_descanso'] or "",
                    reporte['hora_fin_descanso'] or "",
                    reporte['hora_salida'] or "",
                    reporte['observacion'] or ""
                ]
                
                for i, dato in enumerate(datos_individual):
                    texto = str(dato)
                    # Truncar observaciones largas
                    if i == 5 and len(texto) > 18:
                        texto = texto[:18] + "..."
                    p.drawString(col_positions_individual[i], y, texto)
                
                y -= base_row_height
        
        else:
            # REPORTE GENERAL O POR DEPARTAMENTO - Formato completo
            y = height - 100
            base_row_height = 20
            
            # Configuración de columnas para reporte general
            headers = ["ID Usuario", "Nombre", "Departamento", "Fecha", "1ra Marcación", "2da Marcación", "3ra Marcación", "4ta Marcación", "Observaciones"]
            col_positions = [40, 110, 200, 300, 360, 430, 500, 570, 640]
            col_widths = [65, 85, 95, 55, 65, 65, 65, 65, 85]
            
            # Dibujar headers
            p.setFont("Helvetica-Bold", 8)
            for i, header in enumerate(headers):
                p.drawString(col_positions[i], y, header)
            
            # Línea separadora
            p.line(35, y-10, width-35, y-10)
            y -= 20
            
            # Datos con ajuste automático para reporte general
            p.setFont("Helvetica", 8)
            for reporte in reportes:
                # Preparar datos
                datos = [
                    str(reporte['id_usuario']) if reporte['id_usuario'] else "",
                    str(reporte['nombre']) if reporte['nombre'] else "",
                    str(reporte['departamento']) if reporte['departamento'] else "",
                    reporte['fecha'] or "",
                    reporte['hora_ingreso'] or "",
                    reporte['hora_inicio_descanso'] or "",
                    reporte['hora_fin_descanso'] or "",
                    reporte['hora_salida'] or "",
                    str(reporte['observacion']) if reporte['observacion'] else ""
                ]
                
                # Calcular la altura necesaria para esta fila
                max_lines = 1
                wrapped_data = []
                
                for i, dato in enumerate(datos):
                    if dato:
                        # Calcular cuántas líneas necesita este dato
                        chars_per_line = max(1, int(col_widths[i] / 6))
                        wrapped_text = textwrap.fill(str(dato), width=chars_per_line)
                        lines = wrapped_text.split('\n')
                        max_lines = max(max_lines, len(lines))
                        wrapped_data.append(lines)
                    else:
                        wrapped_data.append([''])
                
                row_height = max(base_row_height, max_lines * 12)
                
                # Verificar si necesitamos nueva página
                if y - row_height < 60:
                    p.showPage()
                    y = height - 100
                    # Repetir headers en nueva página
                    p.setFont("Helvetica-Bold", 8)
                    for i, header in enumerate(headers):
                        p.drawString(col_positions[i], y, header)
                    p.line(35, y-10, width-35, y-10)
                    y -= 20
                    p.setFont("Helvetica", 8)
                
                # Resaltar atrasos
                if reporte.get('observacion') == 'Atraso':
                    p.setFillColorRGB(1, 0.95, 0.95)
                    p.rect(35, y-row_height+8, width-70, row_height-3, stroke=0, fill=1)
                    p.setFillColorRGB(0, 0, 0)
                
                # Dibujar cada celda con texto ajustado
                for i, lines in enumerate(wrapped_data):
                    for j, line in enumerate(lines):
                        if line:
                            p.drawString(col_positions[i], y - (j * 10), line)
                
                y -= row_height

    p.save()
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename_pdf}"'
    return response

@login_required
@csrf_exempt
def download_excel_simple(request):
    """Genera y descarga un reporte en Excel simple con formato unificado."""
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'success': False, 'error': 'La librería openpyxl no está instalada.'}, status=500)

    try:
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        id_usuario = request.GET.get('id_usuario')
        nombre = request.GET.get('nombre')
        departamento = request.GET.get('departamento')

        reportes = db_utils.obtener_reportes(fecha_inicio, fecha_fin, id_usuario, nombre, departamento)
        
        # Detectar si es reporte individual
        es_reporte_individual = bool(id_usuario or nombre)

        # Determinar nombre de archivo dinámicamente
        if id_usuario:
            nombre_empleado = reportes[0]['nombre'] if reportes and reportes[0]['nombre'] else f"ID_{id_usuario}"
            filename_xlsx = f"Reporte_de_{nombre_empleado.replace(' ', '_')}.xlsx"
        elif nombre:
            filename_xlsx = f"Reporte_de_{nombre.replace(' ', '_')}.xlsx"
        elif departamento:
            filename_xlsx = f"Reporte_de_{departamento.replace(' ', '_')}.xlsx"
        else:
            filename_xlsx = "Reporte_General.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Asistencia"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        title_font = Font(bold=True, size=14)
        info_font = Font(bold=True, size=11)
        data_font = Font(size=10)
        atraso_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        atraso_font = Font(color="721C24", bold=True)

        current_row = 1

        if es_reporte_individual and reportes:
            # REPORTE INDIVIDUAL - Formato especial
            emp = reportes[0]
            
            # Título principal
            ws.cell(row=current_row, column=1, value=f"REPORTE GENERAL DE {emp['nombre'] or 'EMPLEADO'}").font = title_font
            current_row += 2
            
            # Información del empleado
            ws.cell(row=current_row, column=1, value="INFORMACIÓN DEL EMPLEADO").font = info_font
            current_row += 1
            
            ws.cell(row=current_row, column=1, value="Cédula:").font = info_font
            ws.cell(row=current_row, column=2, value=emp['id_usuario'] or 'No disponible').font = data_font
            current_row += 1
            
            ws.cell(row=current_row, column=1, value="Nombre Completo:").font = info_font
            ws.cell(row=current_row, column=2, value=emp['nombre'] or 'Sin nombre').font = data_font
            current_row += 1
            
            ws.cell(row=current_row, column=1, value="Departamento:").font = info_font
            ws.cell(row=current_row, column=2, value=emp['departamento'] or 'Sin departamento').font = data_font
            current_row += 2
            
            # Período consultado
            ws.cell(row=current_row, column=1, value="PERÍODO CONSULTADO").font = info_font
            current_row += 1
            
            ws.cell(row=current_row, column=1, value="Desde:").font = info_font
            ws.cell(row=current_row, column=2, value=fecha_inicio or 'No especificado').font = data_font
            current_row += 1
            
            ws.cell(row=current_row, column=1, value="Hasta:").font = info_font
            ws.cell(row=current_row, column=2, value=fecha_fin or 'No especificado').font = data_font
            current_row += 3
            
            # Headers simplificados para reporte individual
            headers_individual = ["Fecha", "1ra Marcación", "2da Marcación", "3ra Marcación", "4ta Marcación", "Observaciones"]
            
            for col_num, header_title in enumerate(headers_individual, 1):
                cell = ws.cell(row=current_row, column=col_num, value=header_title)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
            
            # Datos de asistencia para reporte individual
            for reporte in reportes:
                row_data = [
                    reporte.get('fecha'),
                    reporte.get('hora_ingreso'),
                    reporte.get('hora_inicio_descanso'),
                    reporte.get('hora_fin_descanso'),
                    reporte.get('hora_salida'),
                    reporte.get('observacion')
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=value)
                    cell.font = data_font
                
                # Resaltar atrasos
                if reporte.get('observacion') == 'Atraso':
                    for col_num in range(1, len(headers_individual) + 1):
                        cell = ws.cell(row=current_row, column=col_num)
                        cell.fill = atraso_fill
                        cell.font = atraso_font
                
                current_row += 1
        
        else:
            # REPORTE GENERAL O POR DEPARTAMENTO - Formato completo
            # Título
            if departamento:
                ws.cell(row=current_row, column=1, value=f"REPORTE DE {departamento}").font = title_font
            else:
                ws.cell(row=current_row, column=1, value="REPORTE GENERAL").font = title_font
            current_row += 2
            
            # Período consultado
            ws.cell(row=current_row, column=1, value=f"Período: {fecha_inicio or ''} a {fecha_fin or ''}").font = info_font
            current_row += 2
            
            # Headers completos para reporte general
            headers = ["ID Usuario", "Nombre", "Departamento", "Fecha", "1ra Marcación", "2da Marcación", "3ra Marcación", "4ta Marcación", "Observaciones"]
            
            for col_num, header_title in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_num, value=header_title)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
            
            # Agregar datos completos
            for reporte in reportes:
                row_data = [
                    reporte.get('id_usuario'),
                    reporte.get('nombre'),
                    reporte.get('departamento'),
                    reporte.get('fecha'),
                    reporte.get('hora_ingreso'),
                    reporte.get('hora_inicio_descanso'),
                    reporte.get('hora_fin_descanso'),
                    reporte.get('hora_salida'),
                    reporte.get('observacion')
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=value)
                    cell.font = data_font
                
                # Resaltar atrasos
                if reporte.get('observacion') == 'Atraso':
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=current_row, column=col_num)
                        cell.fill = atraso_fill
                        cell.font = atraso_font
                
                current_row += 1

        # Autoajustar columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename_xlsx}"'
        return response

    except Exception as e:
        print(f"Error en download_excel_simple: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@csrf_exempt
def download_excel_advanced(request):
    """Genera y descarga un reporte en Excel avanzado con múltiples hojas."""
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'success': False, 'error': 'La librería openpyxl no está instalada.'}, status=500)

    try:
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        id_usuario = request.GET.get('id_usuario')
        nombre = request.GET.get('nombre')
        departamento = request.GET.get('departamento')

        reportes = db_utils.obtener_reportes(fecha_inicio, fecha_fin, id_usuario, nombre, departamento)

        wb = openpyxl.Workbook()
        
        # Hoja 1: Datos detallados
        ws1 = wb.active
        ws1.title = "Reportes Detallados"
        
        headers = [
            "ID Usuario", "Nombre", "Departamento", "Fecha",
            "1ra Marcación", "2da Marcación", "3ra Marcación", 
            "4ta Marcación", "Observaciones"
        ]
        ws1.append(headers)

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        atraso_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        atraso_font = Font(color="721C24", bold=True)

        for col_num, header_title in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for reporte in reportes:
            row_data = [
                reporte.get('id_usuario'),
                reporte.get('nombre'),
                reporte.get('departamento'),
                reporte.get('fecha'),
                reporte.get('hora_ingreso'),
                reporte.get('hora_inicio_descanso'),
                reporte.get('hora_fin_descanso'),
                reporte.get('hora_salida'),
                reporte.get('observacion')
            ]
            ws1.append(row_data)
            
            if reporte.get('observacion') == 'Atraso':
                for col_num in range(1, len(headers) + 1):
                    cell = ws1.cell(row=ws1.max_row, column=col_num)
                    cell.fill = atraso_fill
                    cell.font = atraso_font

        # Hoja 2: Resumen
        ws2 = wb.create_sheet("Resumen por Departamento")
        dept_summary = {}
        for r in reportes:
            dept = r.get("departamento")
            if dept not in dept_summary:
                dept_summary[dept] = {"total": 0, "atrasos": 0}
            dept_summary[dept]["total"] += 1
            if r.get("observacion") == "Atraso":
                dept_summary[dept]["atrasos"] += 1

        ws2.append(["Departamento", "Total Registros", "Atrasos", "% Puntualidad"])
        for dept, data in dept_summary.items():
            puntualidad = ((data["total"] - data["atrasos"]) / data["total"] * 100) if data["total"] > 0 else 0
            ws2.append([dept, data["total"], data["atrasos"], f"{puntualidad:.1f}%"])

        # Autoajustar columnas en ambas hojas
        for ws in [ws1, ws2]:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename_xlsx = f"reporte_avanzado_{fecha_inicio}_a_{fecha_fin}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename_xlsx}"'
        return response

    except Exception as e:
        print(f"Error en download_excel_advanced: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def api_usuario_info(request, id_usuario):
    """
    API endpoint para obtener información de un usuario específico por ID
    """
    try:
        # Buscar en los reportes recientes para obtener el nombre del usuario
        reportes = db_utils.obtener_reportes(
            fecha_inicio=date.today(),
            fecha_fin=date.today(),
            id_usuario=id_usuario,
            nombre=None,
            departamento=None
        )
        
        if reportes and len(reportes) > 0:
            usuario_info = reportes[0]
            return JsonResponse({
                'success': True,
                'id_usuario': usuario_info.get('id_usuario'),
                'nombre': usuario_info.get('nombre'),
                'departamento': usuario_info.get('departamento')
            })
        else:
            # Si no hay reportes recientes, buscar en un rango más amplio
            from datetime import timedelta
            fecha_hace_30_dias = date.today() - timedelta(days=30)
            reportes_historicos = db_utils.obtener_reportes(
                fecha_inicio=fecha_hace_30_dias,
                fecha_fin=date.today(),
                id_usuario=id_usuario,
                nombre=None,
                departamento=None
            )
            
            if reportes_historicos and len(reportes_historicos) > 0:
                usuario_info = reportes_historicos[0]
                return JsonResponse({
                    'success': True,
                    'id_usuario': usuario_info.get('id_usuario'),
                    'nombre': usuario_info.get('nombre'),
                    'departamento': usuario_info.get('departamento')
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Usuario no encontrado'
                })
                
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def inicio_view(request):
    # Usar la función centralizada para estadísticas avanzadas
    stats = db_utils.obtener_estadisticas_dashboard()
    contexto = {
        'total_usuarios': stats.get('total_empleados', 0),
        'registros_hoy': stats.get('empleados_registrados_hoy', 0),
        'departamentos': db_utils.obtener_total_departamentos(),
        'ausentes_hoy': stats.get('ausentes_hoy', 0),
        'llegadas_tarde': stats.get('llegadas_tarde', 0),
        'porcentaje_asistencia': stats.get('porcentaje_asistencia', 0.0),
        'promedio_horas': stats.get('promedio_horas', '0.0 hrs'),
    }
    return render(request, 'inicio.html', contexto)

from django.contrib.auth import update_session_auth_hash
from .models import Usuario
from .forms import UsuarioCreationForm, UsuarioChangeForm
from django.contrib.auth.forms import UserChangeForm
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
@login_required
def cambiar_credenciales(request):
    if request.method == 'POST':
        nuevo_usuario = request.POST.get('nuevo_usuario')
        nueva_contrasena = request.POST.get('nueva_contrasena')
        confirmar_contrasena = request.POST.get('confirmar_contrasena')
        user = request.user
        errores = []
        if nuevo_usuario:
            if User.objects.filter(username=nuevo_usuario).exclude(pk=user.pk).exists():
                errores.append('El nombre de usuario ya está en uso.')
            else:
                user.username = nuevo_usuario
        if nueva_contrasena:
            if nueva_contrasena != confirmar_contrasena:
                errores.append('Las contraseñas no coinciden.')
            elif len(nueva_contrasena) < 6:
                errores.append('La contraseña debe tener al menos 6 caracteres.')
            else:
                user.set_password(nueva_contrasena)
        if errores:
            for e in errores:
                messages.error(request, e)
        else:
            user.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Credenciales actualizadas correctamente.')
        return redirect('reportes:configuracion_section')
    return redirect('reportes:configuracion_section')

def ayuda_view(request):
    """
    Vista para el Centro de Ayuda.
    """
    contexto = {
        'titulo': 'Centro de Ayuda',
        'descripcion': 'Encuentra respuestas, guías y soporte técnico para el sistema biométrico.'
    }
    return render(request, 'ayuda.html', contexto)
def descargar_excel_registros_hoy(request):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from django.http import HttpResponse
    empleados = db_utils.obtener_empleados()
    usuarios = []
    for emp in empleados:
        ultima = obtener_ultima_marcacion(emp['id_usuario'])
        usuarios.append({
            'id_usuario': emp['id_usuario'],
            'nombre': emp['nombre'],
            'departamento': emp['departamento'],
            'ultima_marcacion': ultima
        })
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros Hoy"
    headers = ["Item", "ID Usuario", "Nombre", "Departamento", "Fecha última marcación", "Hora última marcación"]
    ws.append(headers)
    for idx, usuario in enumerate(usuarios, start=1):
        fecha = usuario['ultima_marcacion']['fecha'] if usuario['ultima_marcacion'] else "-"
        hora = usuario['ultima_marcacion']['hora'] if usuario['ultima_marcacion'] else "-"
        ws.append([
            idx,
            usuario['id_usuario'],
            usuario['nombre'],
            usuario['departamento'],
            fecha,
            hora
        ])
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=registros_hoy.xlsx'
    return response

def descargar_pdf_registros_hoy(request):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, landscape
    from io import BytesIO
    from django.http import HttpResponse
    empleados = db_utils.obtener_empleados()
    usuarios = []
    for emp in empleados:
        ultima = obtener_ultima_marcacion(emp['id_usuario'])
        usuarios.append({
            'id_usuario': emp['id_usuario'],
            'nombre': emp['nombre'],
            'departamento': emp['departamento'],
            'ultima_marcacion': ultima
        })
    buffer = BytesIO()
    page_size = landscape(letter)
    p = canvas.Canvas(buffer, pagesize=page_size)
    width, height = page_size
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, height - 40, "Registros de Hoy")
    y = height - 70
    headers = ["Item", "ID Usuario", "Nombre", "Departamento", "Fecha última marcación", "Hora última marcación"]
    p.setFont("Helvetica-Bold", 10)
    col_positions = [50, 120, 220, 370, 520, 650]
    for i, h in enumerate(headers):
        p.drawString(col_positions[i], y, h)
    y -= 20
    p.setFont("Helvetica", 10)
    for idx, usuario in enumerate(usuarios, start=1):
        fecha = usuario['ultima_marcacion']['fecha'] if usuario['ultima_marcacion'] else "-"
        hora = usuario['ultima_marcacion']['hora'] if usuario['ultima_marcacion'] else "-"
        row = [str(idx), usuario['id_usuario'], usuario['nombre'], usuario['departamento'], str(fecha), str(hora)]
        for i, value in enumerate(row):
            p.drawString(col_positions[i], y, value)
        y -= 20
        if y < 50:
            p.showPage()
            y = height - 70
            p.setFont("Helvetica-Bold", 10)
            for i, h in enumerate(headers):
                p.drawString(col_positions[i], y, h)
            y -= 20
            p.setFont("Helvetica", 10)
    p.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=registros_hoy.pdf'
    return response

from django.contrib.auth.models import User
from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import render, redirect, get_object_or_404

# Solo superusuario puede acceder
superuser_required = user_passes_test(lambda u: u.is_superuser)

from django.contrib.auth.models import User

@superuser_required
def dashboard_admin(request):
    from django.contrib.auth.models import User
    from django.contrib.auth.hashers import make_password
    usuarios = User.objects.all()
    mensaje = None
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        rol = request.POST.get('rol')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        if password1 != password2:
            mensaje = {'tags': 'danger', 'text': 'Las contraseñas no coinciden.'}
        elif User.objects.filter(username=username).exists():
            mensaje = {'tags': 'danger', 'text': 'El nombre de usuario ya existe.'}
        else:
            is_staff = rol == 'staff' or rol == 'superuser'
            is_superuser = rol == 'superuser'
            user = User.objects.create(
                username=username,
                email=email,
                is_staff=is_staff,
                is_superuser=is_superuser,
                password=make_password(password1)
            )
            mensaje = {'tags': 'success', 'text': 'Usuario creado exitosamente.'}
            return redirect('reportes:dashboard_admin')
    return render(request, 'dashboard_admin.html', {'form': None, 'usuarios': usuarios, 'messages': [mensaje] if mensaje else None})

@superuser_required
def usuarios_admin(request):
    usuarios = User.objects.all()
    return render(request, 'usuarios_admin.html', {'usuarios': usuarios})

@superuser_required
def usuario_nuevo(request):
    usuarios = Usuario.objects.all()
    if request.method == 'POST':
        form = UsuarioCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('reportes:usuarios_admin')
    else:
        form = UsuarioCreationForm()
    return render(request, 'usuario_nuevo.html', {'form': form, 'usuarios': usuarios})

@superuser_required
def usuario_editar(request, usuario_id):
    usuario = get_object_or_404(User, pk=usuario_id)
    mensaje = None
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        rol = request.POST.get('rol')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')

        # Validaciones básicas
        if username != usuario.username and User.objects.filter(username=username).exclude(pk=usuario.pk).exists():
            mensaje = {'tags': 'danger', 'text': 'El nombre de usuario ya existe.'}
        elif password1 or password2:
            if password1 != password2:
                mensaje = {'tags': 'danger', 'text': 'Las contraseñas no coinciden.'}
            elif len(password1) < 8:
                mensaje = {'tags': 'danger', 'text': 'La contraseña debe tener al menos 8 caracteres.'}
            else:
                usuario.set_password(password1)
        if not mensaje:
            usuario.username = username
            usuario.email = email
            usuario.is_staff = rol == 'staff' or rol == 'superuser'
            usuario.is_superuser = rol == 'superuser'
            usuario.save()
            mensaje = {'tags': 'success', 'text': 'Usuario actualizado correctamente.'}
            return redirect('reportes:dashboard_admin')
    return render(request, 'usuario_editar.html', {'usuario': usuario, 'messages': [mensaje] if mensaje else None})

@superuser_required
def usuario_eliminar(request, usuario_id):
    usuario = get_object_or_404(User, pk=usuario_id)
    if request.method == 'POST':
        usuario.delete()
        return redirect('reportes:dashboard_admin')
    return render(request, 'usuario_eliminar.html', {'usuario': usuario})
